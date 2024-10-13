from openpyxl import Workbook
from openpyxl.styles import Alignment
from src.lab2.museum.virtual_museum import VirtualMuseum
from src.lab2.reports.report_maker import ReportMaker
import src.lab2.client as client


class ExcelColumnStyle:
    def __init__(self, index: int, header: str, width: int) -> None:
        self.__index = index
        self.__header = header
        self.__width = width

    @property
    def index(self) -> int:
        return self.__index

    @property
    def header(self) -> str:
        return self.__header

    @property
    def width(self) -> int:
        return self.__width


class ExcelReportMaker(ReportMaker):
    __ALIGNMENT = Alignment(horizontal='right', vertical='center')
    __INIT_ROW = 1
    __COLUMNS = {
        'A': ExcelColumnStyle(1, 'Id', 15),
        'B': ExcelColumnStyle(2, 'Название', 30),
        'C': ExcelColumnStyle(3, 'Описание', 40),
        'D': ExcelColumnStyle(4, 'Тип', 15),
        'E': ExcelColumnStyle(5, 'Id коллекции', 20),
        'F': ExcelColumnStyle(6, 'Просмотры пользователя', 25),
        'G': ExcelColumnStyle(7, 'Всего просмотров', 25),
    }

    def __init__(self, museum: VirtualMuseum) -> None:
        super().__init__(museum)
        self.__wb = Workbook()
        self.__ws = self.__wb.active
        self.__row = ExcelReportMaker.__INIT_ROW

    def make(self, path: str = None) -> None:
        self.__preprocess()
        self.__write_data()
        self.__postprocess()
        path = path or f"user_{self._museum.user.id}_{self._museum.user.name}.xlsx"
        self.__wb.save(path)

    def __preprocess(self) -> None:
        self.__ws.title = f"Данные пользователя {self._museum.user.name}"

        for letter, column in ExcelReportMaker.__COLUMNS.items():
            self.__ws.column_dimensions[letter].width = column.width
            self.__ws.cell(self.__row, column.index, column.header)

    def __postprocess(self) -> None:
        for i in range(1, self.__row + 1):
            for letter, column in ExcelReportMaker.__COLUMNS.items():
                self.__ws[f'{letter}{i}'].alignment = ExcelReportMaker.__ALIGNMENT

    def __write_data(self):
        data = client.get_user_info(self._museum.user.id)
        for row in data:
            self.__row += 1
            self.__ws.cell(self.__row, 1, row.get('item_id', None))
            self.__ws.cell(self.__row, 2, row.get('title', None))
            self.__ws.cell(self.__row, 3, row.get('descr', None))
            self.__ws.cell(self.__row, 4, row.get('type', None))
            self.__ws.cell(self.__row, 5, row.get('collection_id', None))
            self.__ws.cell(self.__row, 6, row.get('user_views', None))
            self.__ws.cell(self.__row, 7, row.get('total_views', None))
